# MOSTORO
# PROPER AI

#Importing libraries
import xlwings as xw
from openpyxl.utils import get_column_letter

#Reading schedules
user = input("Enter your laptop user: ")
mec = input("Enter Current Closing Month (Example for March:'03'):")
year = input("Enter Current Year (Example for 2023: '2023'):")


fund1 = xw.Book(r'C:\Users\file 1'.format(user=user, folder=mec,file=mec))
fund2 = xw.Book(r'C:\Users\file 2'.format(user=user, folder=mec,file=mec))
fund3 = xw.Book(r'C:\Users\file 3'.format(user=user, folder=mec,file=mec))
fund4 = xw.Book(r'C:\Users\file 4'.format(user=user, folder=mec,file=mec))
fund5 = xw.Book(r'C:\Users\file 5'.format(user=user, folder=mec,file=mec))
fund6 = xw.Book(r'C:\Users\file 6'.format(user=user, folder=mec,file=mec))
fund7 = xw.Book(r'C:\Users\file 7'.format(user=user, folder=mec,file=mec))
fund8 = xw.Book(r'C:\Users\file 8'.format(user=user, folder=mec,file=mec))
fund9 = xw.Book(r'C:\Users\file 9'.format(user=user, folder=mec,file=mec))

files_to_update = [fund1, fund2, fund3, fund4, fund5, fund6, fund7, fund8, fund9]




for i in range(len(files_to_update)):

    current_schedule = files_to_update[i]
    
    sheet_names = [sheet.name for sheet in current_schedule.sheets]

    #Updating dates
    for prop in range(len(sheet_names)):
        if sheet_names[prop].isdigit():
            current_schedule.sheets['{:}'.format(sheet_names[prop])]['B4'].value = '{mec}/01/{year}'.format(mec=mec, year=year)



    for prop in range(len(sheet_names)):
        if sheet_names[prop].isdigit():

            
            # finding total amount
            row_range = current_schedule.sheets[prop]['A6:z6']
            row_values = row_range.value

            total_index = row_values.index("Current Ending Balance")
            total_cell = row_range[0, total_index]
            last_row = current_schedule.sheets[prop].range(get_column_letter(total_cell.column) + str(current_schedule.sheets[prop].cells.last_cell.row)).end('up').row
            total_amount = current_schedule.sheets[prop][get_column_letter(total_cell.column) + str(last_row)].value
            #print(total_amount)

            # finding beginning balance
            row_range = current_schedule.sheets[prop]['A6:z6']
            row_values = row_range.value
            beg_bal_index = row_values.index("Beginning Balance")
            beg_bal_cell = row_range[0, beg_bal_index]


            last_row = (current_schedule.sheets[prop].range(get_column_letter(beg_bal_cell.column) + str(current_schedule.sheets[prop].cells.last_cell.row)).end('up')).end('up').row
            beg_bal_address = current_schedule.sheets[prop][get_column_letter(beg_bal_cell.column) + str(last_row)].address


            #updating beginning balance amount
            current_schedule.sheets[prop][current_schedule.sheets[prop][get_column_letter(beg_bal_cell.column) + str(last_row)].address].value = total_amount


    print(str(current_schedule),'updated')

